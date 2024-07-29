import datetime
import json
import os
import requests
from requests.auth import HTTPBasicAuth

import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

ROOT_URL = 'https://www.reed.co.uk/api/1.0/search?'
JOB_DETAILS_ROOT = 'https://www.reed.co.uk/api/1.0/jobs/'

class ReedClient:
    
    def __init__(self, api_key: str) -> None:
        self.api_key = api_key
        self.save_directory = "searches"
        os.makedirs(self.save_directory, exist_ok=True)

    def search(self, **args) -> list:
        '''
        Perform a job search with given arguments.
        '''
        return self.process_search_request(ROOT_URL, args)

    def process_search_request(self, url: str, args) -> list:
        '''
        Process job search request using requests library and retrieve all results.
        '''
        auth = HTTPBasicAuth(username=self.api_key, password='')
        args.setdefault('resultsToTake', 100)  # Set default batch size
        jobs = []
        results_fetched = 0

        while True:
            response = requests.get(url, auth=auth, params=args)
            if response.status_code != 200:
                response.raise_for_status()
            
            data = response.json()
            total_results = data['totalResults']
            for job in data['results']:
                if job['minimumSalary'] is None:
                    continue
                del job['employerId']
                del job['employerProfileId']
                del job['employerProfileName']
                jobs.append(job)

            results_fetched += len(jobs)
            print(results_fetched)
            if results_fetched >= total_results:
                break

            args['resultsToSkip'] = results_fetched

        jobs.sort(key=lambda x: datetime.datetime.strptime(x['date'], "%d/%m/%Y"), reverse=True)
        return jobs
    
    def job_details(self, job_id: int) -> dict:
        '''
        Retrieve details of the job with given id.
        '''

        if not type(job_id) is int:
            raise ValueError("'job_id' must be type 'int'")

        url = JOB_DETAILS_ROOT + str(job_id)
        return self.process_job_details_request(url)

    def process_job_details_request(self, url: str) -> dict:
        '''
        Process a job details request using requests library.
        '''
        auth = HTTPBasicAuth(username=self.api_key, password='')
        r = requests.get(url, auth=auth)

        if not r:
            r.raise_for_status()

        return r.json()
    
    def json(self, results, filename) -> None:
        filename = os.path.join(self.save_directory, filename)
        with open(f"{filename}.json","w",encoding="utf8") as file:
            json.dump(results, file, indent=4)

    def excel(self, results, filename) -> None:
        filename = os.path.join(self.save_directory, f"{filename}.xlsx")
        df = pd.DataFrame(results)
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            workbook = writer.book
            worksheet = writer.sheets['Results']
            worksheet.freeze_panes = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            bold_font = Font(bold=True)
            for col in range(1, len(df.columns) + 1):
                cell = worksheet[f"{get_column_letter(col)}1"]
                cell.fill = header_fill
                cell.font = bold_font
                max_length = 0
                column = f'{get_column_letter(col)}:{get_column_letter(col)}'
                for cell in worksheet[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
        os.startfile(filename)