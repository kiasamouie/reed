import pandas as pd
from reed import ReedClient
import os
from urllib.parse import urlencode
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

reed = ReedClient(api_key=r"0da53af5-3512-4e3d-939f-5c773ffb495d")
directory = "searches"
os.makedirs(directory, exist_ok=True)
params = {
    "keywords": "python",
    'minimumSalary': 55000,
}
results = reed.search(**params)
df = pd.DataFrame(results)
excel_filename = os.path.join(directory, urlencode(params).replace('&', '_') + '.xlsx')
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Results')
    workbook = writer.book
    worksheet = writer.sheets['Results']

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)
    for col in range(1, len(df.columns) + 1):
        cell = worksheet[f"{get_column_letter(col)}1"]
        cell.fill = header_fill
        cell.font = bold_font

        max_length = 0
        column = f'{get_column_letter(col)}:{get_column_letter(col)}'
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width

os.startfile(excel_filename)
